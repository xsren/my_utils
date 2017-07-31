# encoding: utf-8
"""
@author: xsren 
@contact: bestrenxs@gmail.com
@site: xsren.me

@version: 1.0
@license: Apache Licence
@file: pack_multiple_excel_to_one.py
@time: 27/07/2017 5:22 PM

将当前目录下的excel合并成一个新的excel
"""
import os
import time
from copy import copy

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.writer.write_only import WriteOnlyCell


def run():
    wb = Workbook()
    ws_w = wb.active
    fpath = '.'

    i = 1
    row_idx = 1
    for fname in os.listdir(fpath):
        if fname.endswith(".xlsx"):
            wb_r = load_workbook(filename=fname)
            ws_r = wb_r.worksheets[0]
            j = 1
            print fname

            # import pdb
            # pdb.set_trace()
            for row in ws_r.rows:

                new_cells = []

                column = 1
                for cell in row:
                    new_cell = WriteOnlyCell(ws_w, value=cell.value)
                    # new_cell = ws_w.cell(row=row_idx,
                    #                      column=column, value=cell.value)
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = copy(cell.number_format)
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)
                    new_cells.append(new_cell)
                    column += 1
                row_idx += 1

                if i == 1:
                    if j == 1:
                        ws_w.title = ws_r.title
                        ws_w.column_dimensions = ws_r.column_dimensions
                    ws_w.append(new_cells)
                    row_idx += 1
                    # ws_w.append(row)
                else:
                    if j >= 4:
                        ws_w.append(new_cells)
                        row_idx += 1
                        # ws_w.append(row)
                j += 1
            i += 1

    today = time.strftime('%Y%m%d_%H:%M:%S', time.localtime(time.time()))
    wb.save('new_file_%s.xlsx' % today)


if __name__ == '__main__':
    run()
